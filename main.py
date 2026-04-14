from fastapi import FastAPI, Depends, HTTPException, Response, Request, UploadFile, File
from auth import get_current_user, require_admin, hash_password, verify_password, create_token
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy.orm import Session
import models, schemas, crud, export
from models import Shipment as ShipmentModel
from database import SessionLocal, engine
import os, logging, io
logging.basicConfig(level=logging.INFO)
models.Base.metadata.create_all(bind=engine)
from database import run_migrations
run_migrations()

from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

app = FastAPI(title="FreightTrack Pro")
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

from fastapi.middleware.cors import CORSMiddleware
_allowed_origins = [o.strip() for o in os.getenv("ALLOWED_ORIGINS", "").split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins or ["http://localhost:8000"],
    allow_credentials=True,
    allow_methods=["GET","POST","PUT","PATCH","DELETE"],
    allow_headers=["Authorization","Content-Type"],
)

app.mount("/static", StaticFiles(directory="static"), name="static")

def get_db():
    db = SessionLocal()
    try: yield db
    finally: db.close()

@app.get("/")
def root():
    from fastapi.responses import HTMLResponse
    import pathlib
    html_path = pathlib.Path("static/index.html")
    content = html_path.read_text(encoding="utf-8") if html_path.exists() else "<h1>Loading...</h1>"
    return HTMLResponse(content=content, headers={
        "Cache-Control": "no-cache, no-store, must-revalidate",
        "Pragma": "no-cache", "Expires": "0", "X-Version": "v27"
    })

@app.get("/track/{ref}")
def client_portal(ref: str):
    return FileResponse("static/portal/index.html")

@app.get("/api/portal/{ref}")
def portal_data(ref: str, db: Session = Depends(get_db)):
    s = crud.get_shipment(db, ref)
    if not s: raise HTTPException(404, "Shipment not found")
    return {
        "ref": s.ref, "ref2": s.ref2, "mode": s.mode, "carrier": s.carrier,
        "vessel": s.vessel, "pol": s.pol, "pod": s.pod, "etd": s.etd, "eta": s.eta,
        "status": s.status, "client": s.client, "lasttracked": s.lasttracked,
        "events": [{"timestamp": e.timestamp, "location": e.location,
                    "description": e.description, "status": e.status}
                   for e in sorted(s.events, key=lambda x: x.timestamp, reverse=True)]
    }

@app.get("/api/shipments", response_model=list[schemas.ShipmentOut])
def list_shipments(q: str = "", search: str = "", status: str = "", mode: str = "",
                   db: Session = Depends(get_db), current=Depends(get_current_user)):
    return crud.get_shipments(db, q or search, status, mode)

@app.post("/api/shipments", response_model=schemas.ShipmentOut)
async def create_shipment(request: Request, db: Session = Depends(get_db),
                          current=Depends(get_current_user)):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON body")
    if "quotationnumber" in body and "quotation_number" not in body:
        body["quotation_number"] = body.pop("quotationnumber")
    ref = (body.get("ref") or "").strip()
    if not ref: raise HTTPException(400, "Reference is required")
    if crud.get_shipment(db, ref): raise HTTPException(400, "Reference already exists")
    allowed = set(schemas.ShipmentCreate.model_fields.keys())
    sdata = {}
    for k, v in body.items():
        if k not in allowed: continue
        sdata[k] = v.strip() if isinstance(v, str) and v.strip() != "" else (None if isinstance(v, str) else v)
    sdata["ref"] = ref
    sdata.setdefault("mode", "Ocean")
    sdata.setdefault("status", "Confirmed")
    try:
        s = schemas.ShipmentCreate(**sdata)
        return crud.create_shipment(db, s)
    except Exception as e:
        import traceback; traceback.print_exc()
        logging.error(f"create_shipment error: {e}"); raise HTTPException(500, "Internal server error")

@app.post("/api/shipments/bulk")
async def bulk_import_shipments(request: Request, db: Session = Depends(get_db),
                                 current=Depends(get_current_user)):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON body")
    rows = body.get("shipments", [])
    if not rows: raise HTTPException(400, "No shipments provided")
    allowed = set(schemas.ShipmentCreate.model_fields.keys())
    created_count = 0
    skipped = []
    for row in rows:
        ref = str(row.get("ref", "")).strip()
        if not ref:
            skipped.append({"ref": None, "reason": "Missing ref"}); continue
        if crud.get_shipment(db, ref):
            skipped.append({"ref": ref, "reason": "Already exists"}); continue
        sdata = {}
        for k, v in row.items():
            if k not in allowed: continue
            sdata[k] = v.strip() if isinstance(v, str) else v
        sdata["ref"] = ref
        sdata.setdefault("mode", "Ocean")
        sdata.setdefault("status", "Confirmed")
        try:
            s = schemas.ShipmentCreate(**sdata)
            crud.create_shipment(db, s)
            created_count += 1
        except Exception as e:
            skipped.append({"ref": ref, "reason": str(e)})
    return {"created": created_count, "skipped": skipped, "total": len(rows)}

@app.get("/api/shipments/{sid}", response_model=schemas.ShipmentOut)
def get_shipment(sid: int, db: Session = Depends(get_db), current=Depends(get_current_user)):
    s = crud.get_shipment_by_id(db, sid)
    if not s: raise HTTPException(404, "Not found")
    return s

@app.put("/api/shipments/{sid}", response_model=schemas.ShipmentOut)
async def update_shipment(sid: int, request: Request, db: Session = Depends(get_db),
                          current=Depends(get_current_user)):
    body = await request.json()
    allowed = set(schemas.ShipmentUpdate.model_fields.keys())
    filtered = {k: (v if v != "" else None) for k, v in body.items() if k in allowed}
    data = schemas.ShipmentUpdate(**filtered)
    s = crud.update_shipment(db, sid, data)
    if not s: raise HTTPException(404, "Not found")
    return s

@app.delete("/api/shipments/{sid}")
def delete_shipment(sid: int, db: Session = Depends(get_db), current=Depends(get_current_user)):
    if current.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    crud.delete_shipment(db, sid)
    return {"ok": True}

@app.patch("/api/shipments/{sid}/status")
async def update_status(sid: int, request: Request, db: Session = Depends(get_db),
                        current=Depends(get_current_user)):
    body = await request.json()
    status = body.get("status", "").strip()
    if not status: raise HTTPException(400, "Status required")
    s = crud.get_shipment_by_id(db, sid)
    if not s: raise HTTPException(404, "Not found")
    s.status = status
    db.commit(); db.refresh(s)
    return s

@app.post("/api/shipments/{sid}/track")
def track_one(sid: int, db: Session = Depends(get_db), current=Depends(get_current_user)):
    import tracker
    s = crud.get_shipment_by_id(db, sid)
    if not s: raise HTTPException(404, "Not found")
    return tracker.track_and_update(db, s)

@app.post("/api/track-all")
def track_all(db: Session = Depends(get_db), current=Depends(get_current_user)):
    import tracker
    for s in crud.get_shipments(db, "", "", ""):
        try: tracker.track_and_update(db, s)
        except: pass
    return {"ok": True}

@app.post("/api/shipments/{sid}/comments", response_model=schemas.CommentOut)
def add_comment(sid: int, data: schemas.CommentCreate, db: Session = Depends(get_db),
                current=Depends(get_current_user)):
    s = crud.get_shipment_by_id(db, sid)
    if not s: raise HTTPException(404, "Not found")
    return crud.add_comment(db, sid, data)

@app.get("/api/export/xlsx")
def export_xlsx(search: str = "", status: str = "", mode: str = "",
                db: Session = Depends(get_db), current=Depends(get_current_user)):
    ships = crud.get_shipments(db, search, status, mode)
    data = export.export_shipments_xlsx(ships)
    return Response(content=data,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=shipments.xlsx"})

@app.get("/api/kpis")
def get_kpis(db: Session = Depends(get_db), current=Depends(get_current_user)):
    return crud.get_kpis(db)

@app.get("/api/stats")
def stats(db: Session = Depends(get_db), current=Depends(get_current_user)):
    return crud.get_stats(db)

@app.get("/api/health")
def health(db: Session = Depends(get_db)):
    try:
        from sqlalchemy import text; db.execute(text("SELECT 1"))
        return {"status": "ok", "db": "connected"}
    except Exception as e:
        return {"status": "error", "db": str(e)}

@app.get("/api/shipments/{sid}/containers")
def get_containers(sid: int, db: Session = Depends(get_db), current=Depends(get_current_user)):
    s = crud.get_shipment_by_id(db, sid)
    if not s: raise HTTPException(404, "Shipment not found")
    return s.containers

@app.post("/api/shipments/{sid}/containers")
def add_container(sid: int, data: dict, db: Session = Depends(get_db),
                  current=Depends(get_current_user)):
    s = crud.get_shipment_by_id(db, sid)
    if not s: raise HTTPException(404, "Shipment not found")
    from models import Container
    cont = Container(shipment_id=sid, container_no=data.get("container_no"),
                     seal_no=data.get("seal_no"), size_type=data.get("size_type"),
                     weight=data.get("weight"))
    db.add(cont); db.commit(); db.refresh(cont)
    return cont

@app.delete("/api/containers/{cid}")
def delete_container(cid: int, db: Session = Depends(get_db), current=Depends(get_current_user)):
    if current.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    from models import Container
    cont = db.query(Container).filter(Container.id == cid).first()
    if not cont: raise HTTPException(404, "Container not found")
    db.delete(cont); db.commit()
    return {"ok": True}

@app.post("/api/shipments/{sid}/send-email")
async def send_email(sid: int, request: Request, db: Session = Depends(get_db),
                     current=Depends(get_current_user)):
    body = await request.json()
    subject = body.get("subject", "").strip()
    email_body = body.get("body", "").strip()
    s = crud.get_shipment_by_id(db, sid)
    if not s: raise HTTPException(404, "Shipment not found")
    if not s.clientemail: raise HTTPException(400, "No client email on this shipment")
    if not subject or not email_body: raise HTTPException(400, "Subject and body required")
    smtp_host = os.getenv("SMTP_HOST", "")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    from_email = os.getenv("FROM_EMAIL", smtp_user)
    if not smtp_host or not smtp_user: raise HTTPException(500, "SMTP not configured")
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject; msg["From"] = from_email; msg["To"] = s.clientemail
        msg.attach(MIMEText(email_body, "plain"))
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls(); server.login(smtp_user, smtp_pass)
            server.sendmail(from_email, s.clientemail, msg.as_string())
        try:
            sent_by = current.get("name", "System") if isinstance(current, dict) else "System"
            crud.log_email(db, sid, subject, s.clientemail, sent_by)
        except: pass
        return {"ok": True, "sent_to": s.clientemail}
    except HTTPException: raise
    except Exception as e: logging.error(f"send_email error: {e}"); raise HTTPException(500, "Internal server error")

@app.get("/api/shipments/{sid}/email-log")
def email_log(sid: int, db: Session = Depends(get_db), current=Depends(get_current_user)):
    try: return crud.get_email_log(db, sid)
    except: return []

@app.post("/api/webhook/shipsgo")
async def webhook(payload: dict, db: Session = Depends(get_db)):
    ref = payload.get("reference") or payload.get("container_number")
    if ref:
        s = crud.get_shipment(db, ref)
        if s:
            import tracker; tracker.track_and_update(db, s)
    return {"ok": True}

import httpx
SHIPSGO_BASE = "https://api.shipsgo.com/v2"

@app.api_route("/proxy/shipsgo/{path:path}", methods=["GET","POST","PATCH","DELETE"])
async def shipsgo_proxy(path: str, request: Request):
    api_key = request.headers.get("X-Shipsgo-User-Token", "")
    body = await request.body()
    params = dict(request.query_params)
    hdrs = {"X-Shipsgo-User-Token": api_key, "Accept": "application/json",
            "Content-Type": "application/json"}
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.request(request.method, f"{SHIPSGO_BASE}/{path}",
                                    headers=hdrs, params=params, content=body)
    try: data = resp.json()
    except: data = resp.text
    return JSONResponse(content=data, status_code=resp.status_code)

@app.get("/debug")
def debug_page(current=Depends(require_admin)): return FileResponse("static/debug.html")

@app.post("/api/shipments/bulk-import")
async def bulk_import(file: UploadFile = File(...), db: Session = Depends(get_db),
                      current=Depends(get_current_user)):
    import openpyxl
    from datetime import datetime
    import re
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    def normalize_header(value):
        raw = str(value or "").strip().lower()
        key = re.sub(r"[^a-z0-9]+", "_", raw).strip("_")
        aliases = {
            "reference":"ref","ref_no":"ref","reference_no":"ref",
            "booking":"bookingno","bookingnumber":"bookingno",
            "clientemail":"clientemail","email":"clientemail",
            "quotation":"quotation_number","quotationno":"quotation_number",
            "lasttracked":"lasttracked","createdat":"createdat",
        }
        return aliases.get(key, key)
    headers = [normalize_header(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    created, skipped, errors = [], [], []
    VALID = ["Confirmed","Booked","Stuffed","Sailing","Arrived","Closed","Canceled"]
    def parse_date(val):
        if not val: return None
        if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
        s = str(val).strip()
        if not s or s.lower() in ('none','null','-','--'): return None
        # Handle timestamp strings like "2026-03-15 00:00:00"
        if len(s) > 10 and s[10] in (' ', 'T'):
            s = s[:10]
        for fmt in ('%Y-%m-%d','%Y/%m/%d','%d/%m/%Y','%d-%m-%Y',
                    '%d.%m.%Y','%m/%d/%Y','%Y%m%d',
                    '%d %b %Y','%d-%b-%Y','%b %d, %Y','%B %d, %Y',
                    '%d %B %Y','%b %d %Y','%B %d %Y'):
            try: return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
            except: pass
        return None
    for row in ws.iter_rows(min_row=2, values_only=True):
        rd = {headers[i]: (str(v).strip() if v is not None else "")
              for i, v in enumerate(row) if i < len(headers)}
        ref = rd.get("ref", "").strip()
        if not ref: continue
        if db.query(models.Shipment).filter(models.Shipment.ref == ref).first():
            skipped.append(ref); continue
        try:
            raw_status = rd.get("status", "").strip()
            teu_raw = rd.get("teu")
            try: teu = int(float(teu_raw)) if teu_raw else None
            except: teu = None
            s = models.Shipment(
                ref=ref,
                ref2=rd.get("ref2", ""),
                booking_no=rd.get("bookingno", "") or rd.get("booking_no", ""),
                mode=rd.get("mode", "Ocean"),
                carrier=rd.get("carrier", ""),
                shipper=rd.get("shipper", ""),
                consignee=rd.get("consignee", ""),
                client=rd.get("client", ""),
                client_email=rd.get("clientemail", "") or rd.get("client_email", ""),
                pol=rd.get("pol", ""),
                pod=rd.get("pod", ""),
                etd=parse_date(rd.get("etd")),
                eta=parse_date(rd.get("eta")),
                status=raw_status if raw_status in VALID else "Confirmed",
                incoterm=rd.get("incoterm", ""),
                vessel=rd.get("vessel", ""),
                voyage=rd.get("voyage", ""),
                teu=teu,
                note=rd.get("note", ""),
                quotation_number=rd.get("quotation_number", ""),
                last_tracked=rd.get("lasttracked", "") or rd.get("last_tracked", ""),
                created_at=datetime.utcnow().isoformat()
            )
            db.add(s); db.commit(); created.append(ref)
        except Exception as e:
            errors.append({"ref": ref, "error": str(e)})
    return {"created": len(created), "skipped": len(skipped), "errors": errors, "refs": created}

@app.post("/api/seed-samples")
def seed_samples(db: Session = Depends(get_db), current=Depends(require_admin)):
    allowed = {c.name for c in models.Shipment.__table__.columns}
    samples = [
        {"ref":"FT-2026-001","mode":"Ocean","carrier":"MSC","client":"Maroc Textiles","pol":"Shanghai","pod":"Casablanca","etd":"2026-01-05","eta":"2026-02-10","status":"Closed","incoterm":"FOB","vessel":"MSC DIANA"},
        {"ref":"FT-2026-002","mode":"Ocean","carrier":"CMA CGM","client":"Atlas Pharma","pol":"Casablanca","pod":"Marseille","etd":"2026-01-12","eta":"2026-01-18","status":"Closed","incoterm":"CIF"},
        {"ref":"FT-2026-003","mode":"Air","carrier":"Royal Air Maroc","client":"TechImport MA","pol":"Hong Kong","pod":"Casablanca","etd":"2026-01-20","eta":"2026-01-21","status":"Closed","incoterm":"EXW"},
        {"ref":"FT-2026-004","mode":"Ocean","carrier":"Maersk","client":"Maroc Textiles","pol":"Casablanca","pod":"Hamburg","etd":"2026-01-25","eta":"2026-02-15","status":"Closed","incoterm":"FOB","vessel":"MAERSK ELBA"},
        {"ref":"FT-2026-005","mode":"Ocean","carrier":"MSC","client":"Casa Ceramics","pol":"Valencia","pod":"Casablanca","etd":"2026-02-01","eta":"2026-02-08","status":"Closed","incoterm":"CFR","vessel":"MSC ANNA"},
    ]
    added = 0
    for s in samples:
        if not db.query(models.Shipment).filter(models.Shipment.ref == s["ref"]).first():
            db.add(models.Shipment(**{k: v for k, v in s.items() if k in allowed}))
            added += 1
    db.commit()
    return {"added": added, "message": f"Seeded {added} sample shipments"}


def _build_legacy_kpi_report(db):
    from collections import defaultdict
    import re

    import_re = re.compile(r'^RO(\d{2})(\d{2})\d+$', re.I)
    export_re = re.compile(r'^ROE(\d{2})(\d{2})\d+$', re.I)

    def ref_parts(ref):
        ref = (ref or '').strip().upper()
        m = export_re.match(ref)
        if m:
            yy, mm = m.group(1), m.group(2)
            return {'dir': 'Export', 'month': f'20{yy}-{mm}'}
        m = import_re.match(ref)
        if m:
            yy, mm = m.group(1), m.group(2)
            return {'dir': 'Import', 'month': f'20{yy}-{mm}'}
        return None

    def norm_dir(sh):
        try:
            d = (getattr(sh, 'direction', None) or '').strip().lower()
            if d in ('export', 'exp', 'x'): return 'Export'
            if d in ('import', 'imp', 'm'): return 'Import'
            p = ref_parts(getattr(sh, 'ref', None))
            return p['dir'] if p else 'Import'
        except Exception:
            return 'Import'

    def norm_mode(sh):
        m = (getattr(sh, 'mode', None) or 'Ocean').strip().lower()
        if m in ('road', 'ftl', 'road ftl', 'truck'): return 'Road'
        if m == 'air': return 'Air'
        return 'Ocean'

    def num(v):
        try: return float(v or 0)
        except: return 0.0

    shipments = db.query(models.Shipment).all()
    total = len(shipments)
    totalteu = round(sum(num(getattr(s, 'teu', 0)) for s in shipments), 2)

    bystatus = defaultdict(int)
    bymode = defaultdict(int)
    bydirection = defaultdict(int)
    monthly_import = defaultdict(int)
    monthly_export = defaultdict(int)
    monthly_teu    = defaultdict(float)
    carrier_all = defaultdict(int)
    carrier_import = defaultdict(int)
    carrier_export = defaultdict(int)
    carrier_road = defaultdict(int)
    carrier_road_import = defaultdict(int)
    carrier_road_export = defaultdict(int)
    incoterm_all = defaultdict(int)
    client_all    = defaultdict(lambda: {'shipments': 0, 'teu': 0.0})
    client_import = defaultdict(lambda: {'shipments': 0, 'teu': 0.0})
    client_export = defaultdict(lambda: {'shipments': 0, 'teu': 0.0})
    client_fcl    = defaultdict(lambda: {'shipments': 0, 'teu': 0.0})
    client_ftl    = defaultdict(lambda: {'shipments': 0, 'teu': 0.0})
    pol_all = defaultdict(int); pol_import = defaultdict(int); pol_export = defaultdict(int)
    pod_all = defaultdict(int); pod_import = defaultdict(int); pod_export = defaultdict(int)
    route_all = defaultdict(int); route_import = defaultdict(int); route_export = defaultdict(int)

    for s in shipments:
        try:
            status = (getattr(s, 'status', None) or 'Pending').strip() or 'Pending'
            direction = norm_dir(s)
            mode = norm_mode(s)
            teu = num(getattr(s, 'teu', 0))
            ref = getattr(s, 'ref', None)
            client = (getattr(s, 'client', None) or '').strip() or None
            carrier = (getattr(s, 'carrier', None) or '').strip() or None
            pol = (getattr(s, 'pol', None) or '').strip().upper()
            pod = (getattr(s, 'pod', None) or '').strip().upper()

            bystatus[status] += 1
            bymode[mode] += 1
            bydirection[direction] += 1

            parts = ref_parts(ref)
            if parts:
                if parts['dir'] == 'Import':
                    monthly_import[parts['month']] += 1
                else:
                    monthly_export[parts['month']] += 1
                monthly_teu[parts['month']] += teu

            if carrier:
                carrier_all[carrier] += 1
                if direction == 'Import': carrier_import[carrier] += 1
                else: carrier_export[carrier] += 1
                if mode == 'Road':
                    carrier_road[carrier] += 1
                    if direction == 'Import': carrier_road_import[carrier] += 1
                    else: carrier_road_export[carrier] += 1
            incoterm = (getattr(s, 'incoterm', None) or '').strip()
            if incoterm:
                incoterm_all[incoterm] += 1

            if client:
                client_all[client]['shipments'] += 1
                client_all[client]['teu'] += teu
                if direction == 'Import':
                    client_import[client]['shipments'] += 1
                    client_import[client]['teu'] += teu
                else:
                    client_export[client]['shipments'] += 1
                    client_export[client]['teu'] += teu
                if mode == 'Ocean':
                    client_fcl[client]['shipments'] += 1
                    client_fcl[client]['teu'] += teu
                else:
                    client_ftl[client]['shipments'] += 1
                    client_ftl[client]['teu'] += teu

            if pol:
                pol_all[pol] += 1
                if direction == 'Import': pol_import[pol] += 1
                else: pol_export[pol] += 1
            if pod:
                pod_all[pod] += 1
                if direction == 'Import': pod_import[pod] += 1
                else: pod_export[pod] += 1
            if pol and pod:
                route = f'{pol} to {pod}'
                route_all[route] += 1
                if direction == 'Import': route_import[route] += 1
                else: route_export[route] += 1
        except Exception:
            continue

    def sort_counts(d, key_name='name', value_name='count', limit=8):
        return [{key_name: k, value_name: v}
                for k, v in sorted(d.items(), key=lambda x: (-x[1], x[0]))[:limit]]

    def sort_client(d, limit=8):
        rows = [{'name': k, 'shipments': int(v['shipments']), 'teu': round(v['teu'], 2)}
                for k, v in d.items()]
        rows.sort(key=lambda x: (-x['shipments'], -x['teu'], x['name']))
        return rows[:limit]

    monthly = []
    for month in sorted(set(monthly_import) | set(monthly_export)):
        monthly.append({
            'month': month,
            'count': monthly_import.get(month, 0) + monthly_export.get(month, 0),
            'import_count': monthly_import.get(month, 0),
            'export_count': monthly_export.get(month, 0),
            'teu': round(monthly_teu.get(month, 0), 1),
        })

    return {
        "total": total,
        "totalteu": totalteu,
        "total_teu": totalteu,
        "bystatus": dict(sorted(bystatus.items())),
        "by_status": dict(sorted(bystatus.items())),
        "bymode": dict(sorted(bymode.items())),
        "by_mode": dict(sorted(bymode.items())),
        "bydirection": {
            "Export": bydirection.get("Export", 0),
            "Import": bydirection.get("Import", 0),
        },
        "by_direction": {
            "Export": bydirection.get("Export", 0),
            "Import": bydirection.get("Import", 0),
        },
        "monthly": monthly,
        "byclientall": sort_client(client_all),
        "by_client_all": sort_client(client_all),
        "byclientimport": sort_client(client_import),
        "by_client_import": sort_client(client_import),
        "by_client_fcl": sort_client(client_fcl),
        "by_client_ftl": sort_client(client_ftl),
        "byclientexport": sort_client(client_export),
        "by_client_export": sort_client(client_export),
        "bycarrier": sort_counts(carrier_all),
        "by_carrier_ocean": sort_counts(carrier_all),
        "by_carrier_road": sort_counts(carrier_road),
        "by_carrier_road_import": sort_counts(carrier_road_import),
        "by_carrier_road_export": sort_counts(carrier_road_export),
        "by_carrier": sort_counts(carrier_all),
        "by_carrier_export": sort_counts(carrier_export),
        "by_carrier_import": sort_counts(carrier_import),
        "bycarrierimport": sort_counts(carrier_import),
        "bycarrierexport": sort_counts(carrier_export),
        "toppol": sort_counts(pol_all),
        "top_pol": sort_counts(pol_all),
        "toppolimport": sort_counts(pol_import),
        "top_pol_import": sort_counts(pol_import),
        "toppolexport": sort_counts(pol_export),
        "top_pol_export": sort_counts(pol_export),
        "toppod": sort_counts(pod_all),
        "top_pod": sort_counts(pod_all),
        "toppodimport": sort_counts(pod_import),
        "top_pod_import": sort_counts(pod_import),
        "toppodexport": sort_counts(pod_export),
        "top_pod_export": sort_counts(pod_export),
        "toprouting": sort_counts(route_all, "route", "count"),
        "top_routing": sort_counts(route_all, "route", "count"),
        "toproutingimport": sort_counts(route_import, "route", "count"),
        "top_routing_import": sort_counts(route_import, "route", "count"),
        "toproutingexport": sort_counts(route_export, "route", "count"),
        "top_routing_export": sort_counts(route_export, "route", "count"),
        "by_incoterm": dict(sorted(incoterm_all.items(), key=lambda x: -x[1])),
        "insights": [],
    }


@app.get("/api/kpi-report")
def kpi_report(db: Session = Depends(get_db), current=Depends(get_current_user)):
    return _build_legacy_kpi_report(db)

@app.get("/api/shipments/{sid}/pdf")
def shipment_pdf(sid: int, db: Session = Depends(get_db), current=Depends(get_current_user)):
    import pdfexport
    s = crud.get_shipment_by_id(db, sid)
    if not s: raise HTTPException(404, "Shipment not found")
    return Response(content=pdfexport.generate_shipment_pdf(s), media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={s.ref}.pdf"})

@app.get("/api/dashboard/pdf")
def dashboard_pdf(db: Session = Depends(get_db), current=Depends(get_current_user)):
    import pdfexport
    from datetime import datetime
    pdf_bytes = pdfexport.generate_dashboard_pdf(crud.get_stats(db), crud.get_shipments(db,"","",""))
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=dashboard_{datetime.now().strftime('%Y%m%d')}.pdf"})

def ensure_admin(db):
    from models import User
    admin_email = os.getenv("ADMIN_EMAIL")
    admin_pass  = os.getenv("ADMIN_PASSWORD")
    if not admin_email or not admin_pass:
        raise RuntimeError(
            "ADMIN_EMAIL and ADMIN_PASSWORD env vars must be set."
        )
    existing = db.query(User).filter(User.role == "admin").first()
    if not existing:
        db.add(User(
            email=admin_email,
            name="Admin", role="admin",
            hashedpw=hash_password(admin_pass),
            isactive=True
        ))
        db.commit()
        print("Default admin created")
    else:
        existing.email    = admin_email
        existing.hashedpw = hash_password(admin_pass)
        db.commit()
        print("Admin credentials synced from env vars")

@app.on_event("startup")
def on_startup():
    from database import Base, engine, SessionLocal, run_migrations
    Base.metadata.create_all(bind=engine)
    run_migrations()
    db = SessionLocal()
    try: ensure_admin(db)
    except Exception as e: print(f"startup: {e}")
    finally: db.close()

@app.post("/api/auth/login")
@limiter.limit("10/minute")
def login(request: Request, body: schemas.LoginRequest, db: Session = Depends(get_db)):
    from models import User
    user = db.query(User).filter(User.email == body.email, User.isactive == True).first()
    if not user or not verify_password(body.password, user.hashedpw):
        raise HTTPException(401, "Invalid email or password")
    token = create_token(user.id, user.role, user.name)
    return {"access_token": token, "role": user.role, "name": user.name}

@app.get("/api/auth/me")
def me(current=Depends(get_current_user)): return current

@app.post("/api/auth/change-password")
def change_password(body: dict, db: Session = Depends(get_db), current=Depends(get_current_user)):
    from models import User
    old_pw = body.get("old_password", ""); new_pw = body.get("new_password", "")
    if not old_pw or not new_pw: raise HTTPException(400, "Both passwords required")
    if len(new_pw) < 10 or not re.search(r'[A-Z]', new_pw) or not re.search(r'[0-9]', new_pw):
        raise HTTPException(400, "Password must be at least 10 characters with 1 uppercase letter and 1 number")
    user = db.query(User).filter(User.id == int(current["sub"])).first()
    if not user or not verify_password(old_pw, user.hashedpw):
        raise HTTPException(401, "Current password is incorrect")
    user.hashedpw = hash_password(new_pw); db.commit()
    return {"message": "Password changed successfully"}

@app.get("/api/users")
def list_users(db: Session = Depends(get_db), current=Depends(require_admin)):
    from models import User
    return [{"id": u.id, "email": u.email, "name": u.name, "role": u.role,
             "is_active": u.isactive, "isactive": u.isactive, "createdat": str(u.createdat)}
            for u in db.query(User).order_by(User.id).all()]

@app.post("/api/users")
def create_user(body: schemas.UserCreate, db: Session = Depends(get_db),
                current=Depends(require_admin)):
    from models import User
    if db.query(User).filter(User.email == body.email).first():
        raise HTTPException(409, "Email already exists")
    u = User(email=body.email, name=body.name, role=body.role,
             hashedpw=hash_password(body.password), isactive=True)
    db.add(u); db.commit(); db.refresh(u)
    return {"id": u.id, "email": u.email, "name": u.name, "role": u.role}

@app.patch("/api/users/{uid}/toggle")
def toggle_user(uid: int, db: Session = Depends(get_db), current=Depends(require_admin)):
    from models import User
    u = db.query(User).filter(User.id == uid).first()
    if not u: raise HTTPException(404, "User not found")
    if u.role == "admin": raise HTTPException(400, "Cannot deactivate admin")
    u.isactive = not u.isactive; db.commit()
    return {"id": u.id, "isactive": u.isactive}

@app.delete("/api/users/{uid}")
def delete_user(uid: int, db: Session = Depends(get_db), current=Depends(require_admin)):
    from models import User
    u = db.query(User).filter(User.id == uid).first()
    if not u: raise HTTPException(404, "User not found")
    if u.role == "admin": raise HTTPException(400, "Cannot delete admin")
    db.delete(u); db.commit()
    return {"deleted": uid}

@app.get("/api/kpi-compare")
def kpi_compare(
    a_from: str = "", a_to: str = "",
    b_from: str = "", b_to: str = "",
    db: Session = Depends(get_db),
    current=Depends(get_current_user)
):
    from collections import defaultdict
    from datetime import datetime
    import calendar, re as _re

    def parse_month(s, end=False):
        if not s: return None
        try:
            dt = datetime.strptime(s.strip()[:7], "%Y-%m")
            if end:
                last = calendar.monthrange(dt.year, dt.month)[1]
                return dt.replace(day=last)
            return dt
        except: return None

    import_re2 = _re.compile(r'^RO(\d{2})(\d{2})\d+$', _re.I)
    export_re2 = _re.compile(r'^ROE(\d{2})(\d{2})\d+$', _re.I)

    def ship_month_dt(s):
        ref = (getattr(s, 'ref', None) or '').strip().upper()
        m = export_re2.match(ref) or import_re2.match(ref)
        if m:
            yy, mm = m.group(1), m.group(2)
            try: return datetime(2000 + int(yy), int(mm), 1)
            except: pass
        etd = getattr(s, 'etd', None)
        if etd:
            try: return datetime.strptime(str(etd)[:10], "%Y-%m-%d")
            except: pass
        return None

    def filter_ships(ships, from_dt, to_dt):
        out = []
        for s in ships:
            dt = ship_month_dt(s)
            if dt is None: continue
            if from_dt and dt < from_dt: continue
            if to_dt and dt.date() > to_dt.date(): continue
            out.append(s)
        return out

    def run_kpi(ships):
        total = len(ships)
        total_teu = round(sum((lambda v: float(v) if v else 0.0)(getattr(s,'teu',None)) for s in ships), 2)
        by_status = defaultdict(int); by_mode = defaultdict(int); by_dir = defaultdict(int)
        monthly_imp = defaultdict(int); monthly_exp = defaultdict(int)
        client_all = defaultdict(lambda: {'shipments':0,'teu':0.0})
        client_fcl = defaultdict(lambda: {'shipments':0,'teu':0.0})
        client_ftl = defaultdict(lambda: {'shipments':0,'teu':0.0})
        carrier_all = defaultdict(int); pol_all = defaultdict(int); pod_all = defaultdict(int)
        pod_exp = defaultdict(int); pod_imp = defaultdict(int); route_all = defaultdict(int)
        now = datetime.utcnow(); overdue = 0
        for s in ships:
            try:
                ref = (getattr(s,'ref',None) or '').strip().upper()
                me = export_re2.match(ref); mi = import_re2.match(ref)
                d = (getattr(s,'direction',None) or '').strip().lower()
                direction = ('Export' if me else 'Import') if (me or mi) else \
                            ('Export' if d in ('export','exp','x') else 'Import')
                status = (getattr(s,'status',None) or 'Pending').strip()
                teu_val = (lambda v: float(v) if v else 0.0)(getattr(s,'teu',None))
                client = (getattr(s,'client',None) or '').strip() or None
                carrier = (getattr(s,'carrier',None) or '').strip() or None
                pol = (getattr(s,'pol',None) or '').strip().upper()
                pod = (getattr(s,'pod',None) or '').strip().upper()
                mode = (getattr(s,'mode',None) or 'Ocean').strip()
                m2 = me or mi
                if m2:
                    month_key = f'20{m2.group(1)}-{m2.group(2)}'
                    if direction == 'Import': monthly_imp[month_key] += 1
                    else: monthly_exp[month_key] += 1
                by_status[status] += 1; by_mode[mode] += 1; by_dir[direction] += 1
                if client:
                    client_all[client]['shipments'] += 1
                    client_all[client]['teu'] += teu_val
                    m_lower = (getattr(s,'mode',None) or 'Ocean').lower()
                    if any(t in m_lower for t in ['ocean','sea','fcl','lcl']):
                        client_fcl[client]['shipments'] += 1
                        client_fcl[client]['teu'] += teu_val
                    else:
                        client_ftl[client]['shipments'] += 1
                        client_ftl[client]['teu'] += teu_val
                if carrier: carrier_all[carrier] += 1
                if pol: pol_all[pol] += 1
                if pod:
                    pod_all[pod] += 1
                    if direction == 'Export': pod_exp[pod] += 1
                    else: pod_imp[pod] += 1
                if pol and pod: route_all[f'{pol} to {pod}'] += 1
                eta = getattr(s,'eta',None)
                if eta and status not in ('Arrived','Closed','Canceled'):
                    try:
                        if now > datetime.strptime(str(eta)[:10], "%Y-%m-%d"): overdue += 1
                    except: pass
            except Exception:
                continue
        def top8(d, key='name', val='count'):
            return [{key:k,val:v} for k,v in sorted(d.items(),key=lambda x:-x[1])[:8]]
        def top_clients(d):
            rows = [{'name':k,'shipments':int(v['shipments']),'teu':round(v['teu'],2)} for k,v in d.items()]
            return sorted(rows,key=lambda x:-x['shipments'])[:8]
        months = sorted(set(list(monthly_imp)+list(monthly_exp)))
        monthly = [{'month':m,'import_count':monthly_imp[m],'export_count':monthly_exp[m],
                    'count':monthly_imp[m]+monthly_exp[m]} for m in months]
        return {
            'total':total,'total_teu':total_teu,
            'by_status':dict(by_status),'by_mode':dict(by_mode),
            'by_direction':{'Export':by_dir['Export'],'Import':by_dir['Import']},
            'monthly':monthly,'by_client_all':top_clients(client_all),
            'by_client_fcl':top_clients(client_fcl),
            'by_client_ftl':top_clients(client_ftl),
            'by_carrier':top8(carrier_all),'top_pol':top8(pol_all),
            'top_pod':top8(pod_all),'top_pod_export':top8(pod_exp),
            'top_pod_import':top8(pod_imp),'top_routing':top8(route_all,key='route'),
            'overdue':overdue,
        }

    all_ships = db.query(models.Shipment).all()
    ships_a = filter_ships(all_ships, parse_month(a_from), parse_month(a_to, end=True))
    ships_b = filter_ships(all_ships, parse_month(b_from), parse_month(b_to, end=True))
    return {
        'period_a': run_kpi(ships_a),
        'period_b': run_kpi(ships_b),
        'a_range': f'{a_from} -> {a_to}',
        'b_range': f'{b_from} -> {b_to}',
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=int(os.getenv("PORT", "8000")))

@app.get("/debug-user-fields")
def debug_user_fields(current=Depends(require_admin)):
    from models import User
    return {"columns": [c.key for c in User.__table__.columns]}
